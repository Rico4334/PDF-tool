from flask import Flask, render_template, request, send_file
import pdfplumber
from openpyxl import load_workbook
import os
from werkzeug.utils import secure_filename

app = Flask(__name__)
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        pdf_file = request.files['pdf']
        filename = secure_filename(pdf_file.filename)
        pdf_path = os.path.join(UPLOAD_FOLDER, filename)
        pdf_file.save(pdf_path)

        data = extract_data_from_pdf(pdf_path)

        output_path = 'Filled_CMR.xlsx'
        fill_excel('template.xlsx', output_path, data)

        return send_file(output_path, as_attachment=True)

    return render_template('index.html')

def extract_data_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        text = page.extract_text()

    lines = text.split('\n')
    load_no = customer_name = destination_address = destination_location = ''
    qty = vol = gw = cases = ''

    for idx, line in enumerate(lines):
        if 'Load No:' in line:
            load_no = line.split(':')[-1].strip()
        if 'Customer Name' in line:
            customer_name = line.split('Customer Name')[-1].strip()
        if 'Final Destination Address' in line:
            destination_address = '\n'.join(lines[idx+1:idx+4])
        if 'Destination Location' in line:
            destination_location = line.split('Destination Location')[-1].strip()
        if 'TOTAL' in line:
            parts = line.split()
            if len(parts) >= 5:
                qty, vol, gw, cases = parts[1:5]

    return {
        'load_no': load_no,
        'customer_name': customer_name,
        'destination_address': destination_address,
        'destination_location': destination_location,
        'totals': {
            'qty': qty,
            'vol': vol,
            'gw': gw,
            'cases': cases
        }
    }

def fill_excel(template_file, output_file, data):
    wb = load_workbook(template_file)
    ws = wb.active

    ws['C20'] = data['load_no']
    ws['C7'] = data['customer_name']
    address_lines = data['destination_address'].split('\n')
    ws['C8'] = address_lines[0] if len(address_lines) > 0 else ''
    ws['C9'] = address_lines[1] if len(address_lines) > 1 else ''
    ws['C10'] = address_lines[2] if len(address_lines) > 2 else ''
    ws['C13'] = data['destination_location']
    ws['I22'] = data['totals']['qty']
    ws['J22'] = data['totals']['vol']
    ws['K22'] = data['totals']['gw']
    ws['L22'] = data['totals']['cases']

    wb.save(output_file)
